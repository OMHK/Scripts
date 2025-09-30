"""
    Search for multiple terms in Excel and return their adjacent values.
    
    Args:
        file_path (str): Path to the Excel file.
        search_terms (list): List of strings to search for.
        offset (tuple): (row_offset, col_offset) → where the value is located relative to the label.
                        Default (0,1) → same row, next column.
    
    Returns:
        dict: {search_term: found_value}
"""
import sys
import os
import logging
import openpyxl
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv


def get_cell_value(ws, row, col):
    """Return the value of a cell, handling merged cells properly."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    # If part of a merged cell, return the top-left value
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            top_left = ws.cell(min_row, min_col)
            return top_left.value
    return None


def find_next_unmerged(ws, row, col):
    """Find the next unmerged cell to the right of (row, col)."""
    for c in range(col + 1, ws.max_column + 1):
        val = get_cell_value(ws, row, c)
        if val is not None or not any(
            (row >= rng.min_row and row <= rng.max_row and c >= rng.min_col and c <= rng.max_col)
            for rng in ws.merged_cells.ranges
        ):
            return row, c
    return row, col


def extract_fields(file_path, search_terms):
    results = {term: None for term in search_terms}
    normalized_terms = {t.lower().strip(): t for t in search_terms}

    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                value = get_cell_value(ws, r, c)

                if value and str(value).strip().lower() in normalized_terms:
                    label = normalized_terms[str(value).strip().lower()]

                    # jump to first unmerged cell after merged label
                    target_row, target_col = find_next_unmerged(ws, r, c)
                    found_value = get_cell_value(ws, target_row, target_col)

                    results[label] = found_value
    return results



# Get file path from command line argument or use default
if len(sys.argv) > 1:
    file_path = sys.argv[1]
else:
    file_path = "input.xlsx"

# Load environment variables
load_dotenv()

# Get fields from environment variables
fields_to_extract = os.getenv('EXCEL_FIELDS', '').split(',')
if not fields_to_extract or fields_to_extract[0] == '':
    fields_to_extract = ["Avancement Financier :", "Montant total des attachements  :", 
                        "Montant du marché après avenants :", "Délai Consommé en jours :"]
    logging.warning("EXCEL_FIELDS not set in .env file, using defaults")

data = extract_fields(file_path, fields_to_extract)

# Create DataFrame and save to CSV
tbl = pd.DataFrame([data])

# Save to CSV with timestamp
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
csv_filename = f'extracted_data_{timestamp}.csv'
tbl.to_csv(csv_filename, index=False, encoding='utf-8')

print(f"Data saved to {csv_filename}")
print("\nExtracted Data:")
print(tbl)
