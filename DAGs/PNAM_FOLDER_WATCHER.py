from datetime import datetime, timedelta
from airflow import DAG
from airflow.sensors.filesystem import FileSensor
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
import os
from datetime import datetime
import json
import openpyxl
import pandas as pd
import logging
from pathlib import Path

# Excel processing functions
def get_cell_value(ws, row, col):
    """Return the value of a cell, handling merged cells properly."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    # If part of a merged cell, return the top-left value
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            top_left = ws.cell(min_row, min_col)
            return top_left.value
    return None

def find_next_unmerged(ws, row, col):
    """Find the next unmerged cell to the right of (row, col)."""
    for c in range(col + 1, ws.max_column + 1):
        val = get_cell_value(ws, row, c)
        if val is not None or not any(
            (row >= rng.min_row and row <= rng.max_row and c >= rng.min_col and c <= rng.max_col)
            for rng in ws.merged_cells.ranges
        ):
            return row, c
    return row, col

def extract_fields(file_path, search_terms):
    """Extract specified fields from Excel file."""
    results = {term: None for term in search_terms}
    normalized_terms = {t.lower().strip(): t for t in search_terms}

    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                value = get_cell_value(ws, r, c)

                if value and str(value).strip().lower() in normalized_terms:
                    label = normalized_terms[str(value).strip().lower()]
                    target_row, target_col = find_next_unmerged(ws, r, c)
                    found_value = get_cell_value(ws, target_row, target_col)
                    results[label] = found_value
    return results

def process_excel_file(file_path, append_mode=True, output_dir=None, **context):
    """Process Excel file and save results to CSV.
    
    This function maintains a single CSV file ('extracted_data.csv') that contains all the extracted data.
    When append_mode is True, new data is added to the existing CSV file.
    When append_mode is False, the CSV file is completely replaced with new data.
    """
    logging.info(f"Processing file: {file_path} with append_mode: {append_mode}")
    
    # Default fields to extract
    fields_to_extract = [
        "Avancement Financier :",
        "Montant total des attachements  :",
        "Montant du marché après avenants :",
        "Délai Consommé en jours :",
        "Semaine du :",
        "au"
    ]

    # Extract data from Excel
    data = extract_fields(file_path, fields_to_extract)
    logging.info(f"Extracted data: {data}")
    
    # Create DataFrame with current file's data
    df = pd.DataFrame([data])
    
    # Add metadata columns
    df['source_file'] = os.path.basename(file_path)
    df['processed_at'] = datetime.now()

    # Set up output directory and file path
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(file_path), 'processed')
    os.makedirs(output_dir, exist_ok=True)
    
    # Single CSV file for all extracted data
    output_file = os.path.join(output_dir, 'extracted_data.csv')
    
    try:
        if append_mode and os.path.exists(output_file):
            # Append to existing CSV file
            existing_df = pd.read_csv(output_file)
            logging.info(f"Existing CSV has {len(existing_df)} records")
            
            # Remove any previous entries for this file if it exists
            existing_df = existing_df[existing_df['source_file'] != os.path.basename(file_path)]
            logging.info(f"After removing previous entries: {len(existing_df)} records")
            
            # Append new data
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            combined_df.to_csv(output_file, index=False, encoding='utf-8')
            logging.info(f"Saved combined CSV with {len(combined_df)} records")
        else:
            # Create new CSV file or replace existing one
            df.to_csv(output_file, index=False, encoding='utf-8')
            logging.info(f"Created new CSV with {len(df)} records")
        
        return output_file
    except Exception as e:
        logging.error(f"Error saving to CSV {output_file}: {str(e)}")
        raise
    
    return output_file

# Function to get file information
def get_file_info(folder_path):
    file_info = {}
    for filename in os.listdir(folder_path):
        filepath = os.path.join(folder_path, filename)
        if os.path.isfile(filepath):
            file_info[filename] = {
                'mtime': os.path.getmtime(filepath),
                'size': os.path.getsize(filepath)
            }
    return file_info

# Function to detect changes
def detect_file_changes(**context):
    folder_path = "/home/omar/Desktop/Test_Directory"
    current_state_file = "/tmp/folder_state.json"
    output_dir = os.path.join(folder_path, 'processed')
    
    # Get current state
    current_state = get_file_info(folder_path)
    
    # Load previous state if exists
    if os.path.exists(current_state_file):
        with open(current_state_file, 'r') as f:
            previous_state = json.load(f)
    else:
        previous_state = {}
    
    # Compare states to find changes
    new_files = []
    modified_files = []
    
    for filename, info in current_state.items():
        if filename.lower().endswith('.xlsx'):  # Only process Excel files
            filepath = os.path.join(folder_path, filename)
            if filename not in previous_state:
                new_files.append(filepath)
            elif info['mtime'] != previous_state[filename]['mtime'] or info['size'] != previous_state[filename]['size']:
                modified_files.append(filepath)
    
    # Initialize the CSV file if it's the first run or if we want to start fresh
    output_file = os.path.join(output_dir, 'extracted_data.csv')
    os.makedirs(output_dir, exist_ok=True)
    
    # Process all files
    processed_files = []
    all_files = new_files + modified_files
    
    if all_files:  # Only process if there are files to process
        # Process first file
        first_file = all_files[0]
        try:
            # First file creates/replaces the CSV if it doesn't exist
            output_file = process_excel_file(
                first_file,
                append_mode=False,  # Replace/create new CSV
                output_dir=output_dir
            )
            processed_files.append(first_file)
            logging.info(f"Successfully processed first file {first_file}")
        except Exception as e:
            logging.error(f"Error processing first file {first_file}: {str(e)}")
        
        # Process remaining files
        for file_path in all_files[1:]:
            try:
                # Subsequent files always append
                output_file = process_excel_file(
                    file_path,
                    append_mode=True,  # Always append for remaining files
                    output_dir=output_dir
                )
                processed_files.append(file_path)
                logging.info(f"Successfully processed {file_path}")
            except Exception as e:
                logging.error(f"Error processing {file_path}: {str(e)}")
    
    # Save current state
    with open(current_state_file, 'w') as f:
        json.dump(current_state, f)
    
    # Push the results to XCom
    context['task_instance'].xcom_push(key='new_files', value=new_files)
    context['task_instance'].xcom_push(key='modified_files', value=modified_files)
    context['task_instance'].xcom_push(key='processed_files', value=processed_files)
    
    return f"Processed files: {processed_files}"

# Default arguments
default_args = {
    "owner": "omar",
    "depends_on_past": False,
    "email_on_failure": False,
    "email_on_retry": False,
    "retries": 1,
    "retry_delay": timedelta(minutes=2),
}

# DAG definition
with DAG(
    "file_monitor_dag",
    default_args=default_args,
    description="Monitor a folder for new Excel files",
    schedule_interval=timedelta(minutes=5),  # check every 5 minutes
    start_date=datetime(2025, 9, 23),
    catchup=False,
    tags=["file", "monitor"],
) as dag:

    # Install required packages if not already installed
    install_packages = BashOperator(
        task_id='install_packages',
        bash_command='pip install openpyxl pandas python-dotenv'
    )

    # 1. Check for file changes and process Excel files
    process_files = PythonOperator(
        task_id='process_excel_files',
        python_callable=detect_file_changes,
        provide_context=True
    )

    # 2. Log processing results
    log_results = BashOperator(
        task_id='log_results',
        bash_command='echo "Processed files: {{ task_instance.xcom_pull(task_ids="process_excel_files", key="processed_files") }}"'
    )

    # Task order
    install_packages >> process_files >> log_results